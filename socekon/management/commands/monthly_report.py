"""
Materiál HR Přehled
* Podle kalendáře na MPSV se vždy stáhnou měsíční statistiky
    nezaměstnanosti z (https://www.mpsv.cz/web/cz/mesicni). V rámci zazipovaných
    dat daného měsíce se pro potřeby HR přehledu bere vždy tabulka č. 4. Zajímá nás
    uchazeči o zaměstnání celkem (bacha nebrat ze sloupce ženy), volná pracovní
    místa celkem a vždy ke konci sled. měsíce a podíl nezaměstnaných osob celkem.
* Uchazeči na 1 VPM se vypočítají vzorcem „uchazeči o zaměstnání děleno volná
    pracovní místa celkem“ (pokud nedojde ke změně zdrojové tabulky, tak stačí do
    té excelové tabulky překopírovat vzorec =J11/X11).
* Obyvatelstvo 15-64 let se bere z MPSV z oddílu
    https://www.mpsv.cz/web/cz/nezamestnanost-v-obcich podle okresů. Stačí vybrat
    nejnovější měsíc, kraj a konkrétní okres a dole opsat hodnotu z řádku celkem.
    Exporty jdou, ale není to s nima úplně jednoduchý. Spíš bych pracoval s těmi
    stránkami (platí obecně, nejen pro tato konkrétní čísla).
* Mzdy se berou z Českého statistického úřadu.
* Rozesílá se mailem na DIZA; STR; DSI; DPP; ZAK
* Aktualizuje se jen na disku K:


"""
from django.core.management.base import BaseCommand, CommandError
from socekon.models import HumanResourcesNuts3, HumanResourcesLau1, Date
from cigeo.models import Lau1, Nuts3
from django.core.exceptions import ObjectDoesNotExist
from pathlib import Path
import json
from openpyxl import load_workbook
import os
from collections import OrderedDict
import copy
import urllib
import time
import sys

from html.parser import HTMLParser
import requests
import tempfile
import atexit
import shutil
import zipfile
import sparql
import ssl
import csv


ssl._create_default_https_context = ssl._create_unverified_context

__FILES__ = []

def __remove_files__():

    global __FILES__
    return
    for f in __FILES__:
        if os.path.isdir(f):
            shutil.rmtree(f)
        else:
            os.remove(f)

atexit.register(__remove_files__)

class StatakHTMLParser(HTMLParser):

    def __init__(self, *args, **kwargs):
        self.tab3_link = None
        super(StatakHTMLParser, self).__init__(*args, **kwargs)

    def handle_starttag(self, tag, attrs):
        if tag == "a":
            for attr in attrs:
                if attr[0] == "href":
                    if attr[1].find("_3.xlsx") > 0:
                        self.tab3_link = attr[1]

class Command(BaseCommand):
    help = 'Import HR data to Django from web services'

    def add_arguments(self, parser):
        parser.add_argument('date', type=str)

    def handle(self, *args, **options):

        nr_nuts3 = 0
        nr_lau1 = 0

        thedate = options["date"]
        (year, month, day) = thedate.split("-")
        #month, year = (int(month), int(year))

        structure = self._get_structure()
        unemployment_data = self._get_unemployment(year, month)
        resources_data = self._get_jobs(unemployment_data["kraje"], year, month)
        data = self.merge(unemployment_data, resources_data)
        wages_nuts3 = self._get_wages_data(year, month)
        population_lau1, population_nuts3 = self._get_population(year, month, structure)

        self.import_data(data, wages_nuts3, population_lau1, population_nuts3, year, month, day)

    @staticmethod
    def _get_structure():
        structure = os.path.join(
                os.path.dirname(os.path.abspath(__file__)), "..",
                "data", "struktura_uzemi_cr_1_1_2016_az_1_1_2020-1.xlsx")

        wb = load_workbook(structure)
        data_sheet = wb.get_sheet_by_name("1.1.2020")
        nuts2 = {}
        nuts3 = {}
        lau1 = {}
        orp = {}
        lau2 = {}

        all_data = list(data_sheet.values)[3:6261]

        for row in all_data:
            (obec_kod, obec_name, status, opu_kod, opu_name, 
            orp_kod, orp_name, lau1_kod, lau1_name, nuts3_kod, nuts3_name, 
            nuts2_kod, nuts2_name) = row

            lau2[int(obec_kod)] = {
                "code": int(obec_kod),
                "name": obec_name,
                "lau1_kod": lau1_kod,
                "lau1_name": lau1_name,
                "nuts3_kod": nuts3_kod,
                "nuts3_name": nuts3_name
            }

        return lau2


    @staticmethod
    def _get_population(year, month, structure):
        sparql_url = "https://data.gov.cz/sparql?query=define%20sql%3Adescribe-mode%20%22CBD%22%20%20DESCRIBE%20%3Chttps%3A%2F%2Fdata.gov.cz%2Fzdroj%2Fdatov%C3%A9-sady%2Fhttp---vdb.czso.cz-pll-eweb-package_show-id-130149%3E&output=application%2Frdf%2Bjson"

        s = requests.get(sparql_url)
        data = s.json()
        link = data[list(data.keys())[0]]["http://www.w3.org/ns/dcat#distribution"]
        download_sparql = link[0]["value"]
        
        sparql_url = "https://data.gov.cz/sparql?query=define%20sql%3Adescribe-mode%20%22CBD%22%20%20DESCRIBE%20%3C{}%3E&output=application%2Frdf%2Bjson".format(download_sparql)

        s = requests.get(sparql_url)
        data = s.json()
        link = data[list(data.keys())[0]]["http://www.w3.org/ns/dcat#downloadURL"]
        zip_link = link[0]["value"]

        fd, zip_file_name = tempfile.mkstemp("zip-population")
        __FILES__.append(zip_file_name)

        z = requests.get(zip_link)


        thedir = tempfile.mkdtemp(prefix="population-")
        __FILES__.append(thedir)

        population_zip = os.path.join(thedir, "{}-{}.zip".format(year, month))

        with open(population_zip, "wb") as zip_file:
            zip_file.write(z.content)

        zipfile_class = zipfile.ZipFile(population_zip)
        files = zipfile_class.namelist()
        population_csv = os.path.join(thedir, files[0])
        zipfile_class.extract(os.path.basename(population_csv), path=thedir)

        assert os.path.isfile(population_csv)

        population_data =  {}
        with open(population_csv) as csvfile:
            dialect = csv.Sniffer().sniff(csvfile.read(2048))
            csvfile.seek(0)
            reader = csv.reader(csvfile)
            for row in reader:

                (idhod, hodnota, stapro_kod, pohlavi_cis, pohlavi_kod,
                vuzemi_cis, vuzemi_kod, rok, casref_do, pohlavi_txt,
                vuzemi_txt) = row

                if idhod == "idhod":
                    continue

                (pyear, pmonth, pday) = (int(v) for v in casref_do.split("-"))
                rok = int(rok)
                vuzemi_kod = int(vuzemi_kod)

                if rok not in population_data:
                    population_data[rok] = {}
                #    if vuzemi_kod not in population_data[rok]:
                #        population_data[rok][vuzemi_kod] = 0

                if not pohlavi_kod:
                    population_data[rok][vuzemi_kod] = int(hodnota)

        year = int(year)
        if year in population_data:
            year_data = population_data[year]
        elif year - 1 in population_data:
            year_data = population_data[year-1]
        elif year - 2 in population_data:
            year_data = population_data[year-2]


        missing = []

        population_lau1 = {}
        population_nuts3 = {}

        for k in year_data.keys():
            if int(k) in structure:
                k = int(k)
                lau1_name = structure[k]["lau1_name"]
                nuts3_name = structure[k]["nuts3_name"]
                if lau1_name == "Praha":
                    lau1_name = "Hlavní město Praha"

                if lau1_name not in population_lau1:
                    population_lau1[lau1_name] = 0

                if nuts3_name not in population_nuts3:
                    population_nuts3[nuts3_name] = 0

                population_lau1[lau1_name] += year_data[k]
                population_nuts3[nuts3_name] += year_data[k]

            else:
                missing.append((k, year_data[k]))

        return (population_lau1, population_nuts3)


    def import_lau1(self, data, population,  year, month, day):
        for lau1_name in data:
            lau1 = Lau1.objects.get(name=lau1_name)
            date = Date.objects.get(date="{}-{}-{}".format(year, month, day))
            inhabitans = population[lau1_name]
            productive_inhabitans = data[lau1_name]["productive_inhabitans"]
            unemployed = data[lau1_name]["unemployed"]
            vacancies = data[lau1_name]["vacancies"]
            unemployment = data[lau1_name]["unemployment"]
            applications_per_vacancy = data[lau1_name]["applications_per_vacancy"]
            if len(HumanResourcesLau1.objects.filter(lau1=lau1, date=date)):
                self.stdout.write(self.style.WARNING("LAU1 for {} and {} exists already, remove first".format(lau1.name, date.date)))

            else:
                HumanResourcesLau1.objects.create(
                    lau1=lau1,
                    date=date,
                    inhabitans=inhabitans,
                    productive_inhabitans=productive_inhabitans,
                    unemployed=unemployed,
                    vacancies=vacancies,
                    unemployment=unemployment,
                    applications_per_vacancy=applications_per_vacancy
                )
                self.stdout.write(self.style.SUCCESS('Successfully imported {}'.format(lau1.name)))

    def import_nuts3(self, data, wages, population,  year, month, day):
        for nuts3_name in data:
            nuts3 = Nuts3.objects.get(name=nuts3_name)
            date = Date.objects.get(date="{}-{}-{}".format(year, month, day))
            nuts3_wages = wages[nuts3_name]
            inhabitans = population[nuts3_name]
            productive_inhabitans = data[nuts3_name]["productive_inhabitans"]
            unemployed = data[nuts3_name]["unemployed"]
            vacancies = data[nuts3_name]["vacancies"]
            unemployment = data[nuts3_name]["unemployment"]
            applications_per_vacancy = data[nuts3_name]["applications_per_vacancy"]
            if len(HumanResourcesNuts3.objects.filter(nuts3=nuts3, date=date)):
                self.stdout.write(self.style.WARNING(
                    "NUTS3 for {} and {} exists already, remove first".format(nuts3.name, date.date))
                    )

            else:
                HumanResourcesNuts3.objects.create(
                    nuts3=nuts3,
                    date=date,
                    wages=nuts3_wages,
                    inhabitans=inhabitans,
                    productive_inhabitans=productive_inhabitans,
                    unemployed=unemployed,
                    vacancies=vacancies,
                    unemployment=unemployment,
                    applications_per_vacancy=applications_per_vacancy
                )
                self.stdout.write(self.style.SUCCESS('Successfully imported {}'.format(nuts3.name)))


    def import_data(self, data, wages_nuts3, population_lau1, population_nuts3, year,
            month, day):

        self.import_lau1(data["okresy"],  population_lau1, year, month, day)
        self.import_nuts3(data["kraje"],  wages_nuts3, population_nuts3, year, month, day)


    @staticmethod
    def _get_wages_data(year, month):
        global __FILES__
        month = int(month)

        q = 1
        if month <= 3:
            q = 1
        elif month <= 6:
            q = 2
        elif month <= 9:
            q = 3
        elif month <= 12:
            q = 4

        parser = StatakHTMLParser()

        while not parser.tab3_link:
            link = "https://statistika.info/csu/czso/cri/prumerne-mzdy-{}-ctvrtleti-{}".format(q, year)
            response = requests.get(link, verify=False)
            parser.feed(response.content.decode("utf-8"))

            if not parser.tab3_link:
                q = q - 1
                if q < 1:
                    q = 4
                    year  = int(year) - 1

        assert parser.tab3_link

        (fd, wages_quartal_filename) = tempfile.mkstemp(prefix="wages-", suffix=".xlsx")
        __FILES__.append(wages_quartal_filename)

        response = requests.get(parser.tab3_link)
        with open(wages_quartal_filename, "wb") as out:
            out.write(response.content)

        wb = load_workbook(wages_quartal_filename)
        data_sheet = wb.get_sheet_by_name("List1")
        all_data = list(data_sheet.values)[9:31]
        data = {}

        for row in all_data:
            if row[0] == "Hl. m. Praha":
                data["Hlavní město Praha"] = row[5]
            else:
                data[row[0]] = row[5]
        return data

    def _merge_type(self, unempl, resource, mytype):

        data = {}
        for name in unempl[mytype]:
            if name in resource[mytype]:
                data[name] = OrderedDict(
                    #inhabitans=
                    name=name,
                    productive_inhabitans=resource[mytype][name]["obyvatelstvo15_64"],
                    unemployed=unempl[mytype][name]["unemployed"],
                    vacancies=unempl[mytype][name]["vacancies"],
                    unemployment=unempl[mytype][name]["unemployment"],
                    applications_per_vacancy=unempl[mytype][name]["applications_per_vacancy"]
                )
        return data

    def merge(self, unempl, resource):
        okresy = self._merge_type(unempl, resource, "okresy")
        kraje = self._merge_type(unempl, resource, "kraje")
        return {"kraje": kraje, "okresy": okresy}


    def _get_unemployment(self, year, month):
        global __FILES__
        response = requests.get("https://www.mpsv.cz/o/rest/statistiky/nezamestnanost/{}/{}".format(year,
            month))

        thedir = tempfile.mkdtemp(prefix="unemployment-")
        __FILES__.append(thedir)

        unemployment_zip = os.path.join(thedir, "{}-{}.zip".format(year, month))
        unemployment_excel =  os.path.join(thedir, "4. NEZ{}{}h.xlsx".format(month, int(year)-2000))

        with open(unemployment_zip, "bw") as out_zip:
            out_zip.write(response.content)

        zipfile_class = zipfile.ZipFile(unemployment_zip)

        for f in ("4. NEZ{}{}h.xlsx", "4. Nez{}{}h.xlsx", "4. nez{}{}h.xlsx"):
            try:
                unemployment_excel =  os.path.join(thedir, f.format(month, int(year)-2000))
                zipfile_class.extract(os.path.basename(unemployment_excel), path=thedir)
                break
            except KeyError as e:
                pass

        assert os.path.isfile(unemployment_excel)

        wb = load_workbook(unemployment_excel)
        data_sheet = wb.get_sheet_by_name("nuts3")
        all_data = list(data_sheet.values)[10:100]
        data = {"okresy":{}, "kraje":{}}

        okresy_in_kraj = []
        for row in all_data:
            name = row[1]
            unemployed = row[9]
            vacancies = row[23]
            unemployment = row[29]
            applications_per_vacancy = unemployed / vacancies

            record = {
                'name': name,
                'unemployed': unemployed,
                'vacancies': vacancies,
                'unemployment': unemployment,
                'applications_per_vacancy': applications_per_vacancy
            }

            if name == "Praha":

                data["kraje"]["Hlavní město Praha"] = record
                data["kraje"]["Hlavní město Praha"]["okresy"] = ["Hlavní město Praha"]
                data["okresy"]["Hlavní město Praha"] = record
                continue

            if name.lower().find("kraj") > -1:
                data["kraje"][name] = record
                data["kraje"][name]["okresy"] = copy.copy(okresy_in_kraj)
                okresy_in_kraj = []
            else:
                okresy_in_kraj.append(name)
                data["okresy"][name] = record
        return data



    def _get_jobs(self, kraje, year, month):

        data = (
            "PREFIX zdroj: <https://data.mpsv.cz/zdroj/>\n"
            "PREFIX pojem: <https://data.mpsv.cz/pojem/>\n"
            "PREFIX skos: <http://www.w3.org/2004/02/skos/core#>\n"
            "SELECT \n"
            "SUM(?uchazeciOZamestnani) AS ?uchazeciOZamestnani\n"
            "SUM(?dosazitelniUchazeci) AS ?dosazitelniUchazeci\n"
            "SUM(?podilNezamestnanychOsob) AS ?podilNezamestnanychOsob\n"
            "SUM(?volnaPracovnaMista) AS ?volnaPracovnaMista\n"
            "SUM(?obyvatelstvo15_64) AS ?obyvatelstvo15_64\n"
            "?nazevObce\n"
            "?nazevOkresu\n"
            "?rok\n"
            "?mesic\n"
            "FROM zdroj:NezamestnanostVObcich\n"
            "FROM NAMED zdroj:Okresy\n"
            "FROM NAMED zdroj:Kraje\n"
            "FROM NAMED zdroj:Obce\n"
            "WHERE \n"
            "{{\n"
            "?polozka pojem:mesic ?mesic .\n"
            "FILTER (MONTH(?mesic) = {month})\n"
            "?polozka pojem:rok ?rok .\n"
            "FILTER (YEAR(?rok) = {year})\n"
            "?polozka pojem:okres ?idOkresu .\n"
            "?polozka pojem:obec ?idObce .\n"
            "?polozka pojem:dosazitelniUchazeci ?dosazitelniUchazeci .\n"
            "?polozka pojem:podilNezamestnanychOsob ?podilNezamestnanychOsob .\n"
            "?polozka pojem:uchazeciOZamestnani ?uchazeciOZamestnani .\n"
            "?polozka pojem:volnaPracovnaMista ?volnaPracovnaMista .\n"
            "?polozka pojem:obyvatelstvo15_64 ?obyvatelstvo15_64\n"
            "GRAPH zdroj:Okresy\n"
            "{{\n"
            "?idOkresu skos:prefLabel ?nazevOkresu .\n"
            "?idOkresu pojem:kraj ?idKraje\n"
            "}}\n"
            "GRAPH zdroj:Kraje\n"
            "{{\n"
            "?idKraje skos:prefLabel ?nazevKraje\n"
            "}}\n"
            "GRAPH zdroj:Obce\n"
            "{{\n"
            "?idObce skos:prefLabel ?nazevObce\n"
            "}}\n"
            "}}\n"
            "GROUP BY ?rok ?mesic ?nazevObce ?nazevOkresu\n"
            "ORDER BY DESC(?rok), DESC(?mesic), ASC(?nazevObce), ASC(?nazevOkresu)"
            ).format(month=month, year=year)
        s = sparql.Service("https://www.mpsv.cz/sparql/", "utf-8", "POST",
                           accept='application/json')
        while True:
            try:
                result = s.query(data, raw=True)
                break
            except urllib.error.HTTPError as e:
                print("Sparql failed, re-trying")
                time.sleep(5)
        #response = requests.post("https://www.mpsv.cz/sparql/", data=data,
        #        headers={"Accept": "application/json"})
        data = json.load(result)
        head = data["head"]
        data = data["results"]

        okresy = {}
        newkraje = {}
        for obec in data["bindings"]:

            nazevOkresu = obec["nazevOkresu"]["value"]
            if nazevOkresu == "Plzeň":
                nazevOkresu = "Plzeň-město"
            if nazevOkresu == "Ostrava":
                nazevOkresu = "Ostrava-město"
            #if nazevOkresu == "Hlavní město Praha":
            #    nazevOkresu = "Praha"
            kraj = self._get_kraj_name(nazevOkresu, kraje)
            assert kraj

            if not nazevOkresu in okresy:
                okresy[nazevOkresu] = {}
                okresy[nazevOkresu]["uchazeciOZamestnani"] = 0
                okresy[nazevOkresu]["dosazitelniUchazeci"] = 0
                okresy[nazevOkresu]["volnaPracovnaMista"] = 0
                okresy[nazevOkresu]["obyvatelstvo15_64"] = 0

            if not kraj in newkraje:
                newkraje[kraj] = {}
                newkraje[kraj]["uchazeciOZamestnani"] = 0
                newkraje[kraj]["dosazitelniUchazeci"] = 0
                newkraje[kraj]["volnaPracovnaMista"] = 0
                newkraje[kraj]["obyvatelstvo15_64"] = 0

            okresy[nazevOkresu]["uchazeciOZamestnani"] += int(obec["uchazeciOZamestnani"]["value"])
            okresy[nazevOkresu]["dosazitelniUchazeci"] += int(obec["dosazitelniUchazeci"]["value"])
            okresy[nazevOkresu]["volnaPracovnaMista"] += int(obec["volnaPracovnaMista"]["value"])
            okresy[nazevOkresu]["obyvatelstvo15_64"] += int(obec["obyvatelstvo15_64"]["value"])

            newkraje[kraj]["uchazeciOZamestnani"] += int(obec["uchazeciOZamestnani"]["value"])
            newkraje[kraj]["dosazitelniUchazeci"] += int(obec["dosazitelniUchazeci"]["value"])
            newkraje[kraj]["volnaPracovnaMista"] += int(obec["volnaPracovnaMista"]["value"])
            newkraje[kraj]["obyvatelstvo15_64"] += int(obec["obyvatelstvo15_64"]["value"])


        return {"okresy": okresy, "kraje": newkraje}

    def _get_kraj_name(self, okres, kraje):

        for kraj in kraje:
            if okres in kraje[kraj]["okresy"]:
                return kraj
